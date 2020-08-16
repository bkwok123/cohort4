import pytest
import os
import datetime
from pathlib import Path
from openpyxl import load_workbook
from testfixtures import TempDirectory
from generateExInv import generateExInv, validateInvInput, loadWStoDictionary, validateDictionary, checkType, checkPostalCode, checkPhone, printError, createInvoice, inner_join, merge_input_template

@pytest.fixture
def dirpath():
    return Path("C:/code/cohort4/python-Excel/template")
@pytest.fixture    
def outfilePrefix():
    return "Excel Invoice"
@pytest.fixture    
def infile():
    return "Merge_template.xlsx"

@pytest.fixture(params=[
        {"fields": ["customer_id", "first_name", "last_name", "phone", "address", "city", "province", "postal_code"],
         "wsname": "Customer"},
        {"fields": ["invoice_id", "customer_id", "invoice_date"],
         "wsname": "Invoice"},         
        {"fields": ["invoice_line_Item_id", "invoice_id", "product_id", "item_ref", "quantity"],
         "wsname": "Invoice Line Item"},                  
        {"fields": ["product_id", "name", "description", "unit_price"],
         "wsname": "Product"},                           
    ])

def setup(request):
    retVal = request.param
    # print("\nSetup! retVal = {}".format(retVal))
    return retVal

#################################################################
# Exercise - Populate the data
#################################################################    
def test_generateExInv(dirpath, outfilePrefix):    

    data = {1: {'invoice_id': 1, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 4, 18, 0, 0), 
            'first_name': 'John', 'last_name': 'Doe', 'phone': 4031234567, 'address': '123 Fake Street', 'city': 'Calgary', 'province': 'AB', 'postal_code': 'T1X1N1', 
            'invoice_line_Item_id': 1, 'product_id': 1, 'item_ref': 'Item 1', 'quantity': 3, 'name': 'Pen', 'description': 'Ball Pointed, Black Ink', 'unit_price': 3}, 
            2: {'invoice_id': 1, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 4, 18, 0, 0), 
            'first_name': 'John', 'last_name': 'Doe', 'phone': 4031234567, 'address': '123 Fake Street', 'city': 'Calgary', 'province': 'AB', 'postal_code': 'T1X1N1', 
            'invoice_line_Item_id': 2, 'product_id': 2, 'item_ref': 'Item 2', 'quantity': 1, 'name': 'Pencil', 'description': 'Mechanical, 0.3mm', 'unit_price': 5}, 
            3: {'invoice_id': 1, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 4, 18, 0, 0), 
            'first_name': 'John', 'last_name': 'Doe', 'phone': 4031234567, 'address': '123 Fake Street', 'city': 'Calgary', 'province': 'AB', 'postal_code': 'T1X1N1', 
            'invoice_line_Item_id': 3, 'product_id': 3, 'item_ref': 'Item 3', 'quantity': 1, 'name': 'Eraser', 'description': 'White', 'unit_price': 2}}

    outfile = f"{outfilePrefix} 1.xlsx"

    with TempDirectory() as d:      
        create_status = generateExInv(d.path,outfilePrefix,data)        
        filename = os.path.join(d.path, outfile)
        existEX = os.path.exists(filename)
        wb = load_workbook(filename)
        ws = wb.active       
        d.listdir()      

    assert existEX == True
    assert create_status == True
    assert ws['A1'].value == "Invoice Date:"
    assert ws['B1'].value == datetime.datetime(2020, 4, 18, 0, 0)
    assert ws['A2'].value == "Invoice Number:"  
    assert ws['B2'].value == 1
    assert ws['E1'].value == "Billed to:"
    assert ws['F1'].value == "John Doe"

    assert ws['A9'].value == "Item 1"
    assert ws['B9'].value == "Pen"
    assert ws['C9'].value == "Ball Pointed, Black Ink"

#################################################################
# Exercise - Validate the data
#################################################################
def test_validateInvInput(monkeypatch, capsys, dirpath, infile):

    filename = os.path.join(dirpath, infile)
    
    ws1_exist = False
    ws2_exist = False
    ws3_exist = False
    ws4_exist = False    

    input_exist = os.path.isfile(filename)
    if input_exist:
        wb = load_workbook(filename)
        if "Customer" in wb.sheetnames:
            ws1_exist = True
        if "Invoice" in wb.sheetnames:
            ws2_exist = True 
        if "Invoice Line Item" in wb.sheetnames:        
            ws3_exist = True
        if "Product" in wb.sheetnames:
            ws4_exist = True               

    wbDict = validateInvInput(dirpath, infile)
    valid_status = wbDict["Validation"]
    missing_ws = ws1_exist & ws2_exist & ws3_exist & ws4_exist
    
    captured = capsys.readouterr()

    assert "WB" in wbDict.keys()
    assert "Error" in wbDict.keys()
    assert "Validation" in wbDict.keys()
    assert "Customer" in wbDict["WB"]
    assert "Invoice" in wbDict["WB"]
    assert "Invoice Line Item" in wbDict["WB"]
    assert "Product" in wbDict["WB"]

    if (valid_status == False):
        assert "Error Count: " in captured.out
    if (missing_ws == False):
        assert "Missing worksheets: " in captured.out
        if (ws1_exist == False):
            assert "Customer" in captured.out
        if (ws2_exist == False):
            assert "Invoice" in captured.out    
        if (ws3_exist == False):
            assert "Invoice Line Item" in captured.out
        if (ws4_exist == False):
            assert "Product" in captured.out

def test_loadWStoDictionary(dirpath, infile, setup):
    
    filename = os.path.join(dirpath, infile)
    input_exist = os.path.isfile(filename)
    if input_exist:
        wb = load_workbook(filename) 
        ws = wb[setup["wsname"]]
        dictionary = loadWStoDictionary(ws, setup["fields"])
        print(dictionary)

        for field in setup["fields"]:
            assert field in dictionary[1].keys()

def test_checkType():    
    assert checkType(123, int) == ""                            
    assert checkType("123", int) == "Incorrect Data Type, Expected: <class 'int'>, Received: <class 'str'>\n"
    assert checkType("123", str) == ""                            
    assert checkType("ABC", str) == ""                            
    assert checkType("ABC", int) == "Incorrect Data Type, Expected: <class 'int'>, Received: <class 'str'>\n"
    assert checkType(12.3, float) == ""
    assert checkType(12, float) == ""
    assert checkType(12.0, float) == ""
    assert checkType("12.3", float) == "Incorrect Data Type, Expected: <class 'float'>, Received: <class 'str'>\n"

def test_checkPostalCode():    
    assert checkPostalCode("123") == "Incorrect Postal Format"  
    assert checkPostalCode("123567") == "Incorrect Postal Format"  
    assert checkPostalCode("1a3a6a") == "Incorrect Postal Format"  
    assert checkPostalCode("a2a5a7") == ""  
    assert checkPostalCode("a2a 5a7") == ""

def test_checkPhone():
    assert checkPhone("4031234567") == ""      
    assert checkPhone("403 -123 -4567") == ""
    assert checkPhone("(403)123-4567") == ""
    assert checkPhone("(403)123-45678") == "Incorrect Phone Format"
    assert checkPhone("(403)123-45A7") == "Incorrect Phone Format"
    assert checkPhone("(403)abc-45A7") == "Incorrect Phone Format"

@pytest.mark.parametrize(
    "fields,dictionary,expected",
    [({"customer_id": {"func": [], "type": int},
       "first_name": {"func": [], "type": str},
       "last_name": {"func": [], "type": str},
       "phone": {"func": [checkPhone], "type": str},
       "address": {"func": [], "type": str},
       "city": {"func": [], "type": str},
       "province": {"func": [], "type": str},
       "postal_code": {"func": [checkPostalCode], "type": str}},
      {1: {'customer_id': 1, 'first_name': 'John', 'last_name': 'Doe', 'phone': 4031234567, 'address': '123 Fake Street', 'city': 'Calgary', 'province': 'AB', 'postal_code': 'T1X1N1'}, 
       2: {'customer_id': 2, 'first_name': 'Jane', 'last_name': None, 'phone': 7081234567, 'address': '456 Fake Avenue', 'city': 'Edmonton', 'province': 'AB', 'postal_code': 'D1Z1X1'}, 
       3: {'customer_id': 3, 'first_name': 'Noname', 'last_name': 'OK', 'phone': 7081234567, 'address': '456 Fake Avenue', 'city': 'Edmonton', 'province': 'AB', 'postal_code': 'D1Z1X1'}},
      {'rowErrorMsg': {2: 'last_name:\n Empty Value\n'}, 'MissingField': {}, 'errorCount': 1}),

     ({"customer_id": {"func": [], "type": int},
       "first_name": {"func": [], "type": str},
       "last_name": {"func": [], "type": str},
       "phone": {"func": [checkPhone], "type": str},
       "address": {"func": [], "type": str},
       "city": {"func": [], "type": str},
       "province": {"func": [], "type": str},
       "postal_code": {"func": [checkPostalCode], "type": str}},
      {1: {'customer_id': "a", 'first_name': 'John', 'last_name': 'Doe', 'phone': 4031234567, 'address': '123 Fake Street', 'province': 'AB'}, 
       2: {'customer_id': 2, 'first_name': 'Jane', 'last_name': None, 'phone': 7081234567, 'address': '456 Fake Avenue', 'province': 'AB'}, 
       3: {'customer_id': 3, 'first_name': 'Noname', 'last_name': 'OK', 'phone': 7081234567, 'address': '456 Fake Avenue', 'province': 'AB'}},       
      {'rowErrorMsg': {1: "customer_id:\n Incorrect Data Type, Expected: <class 'int'>, Received: <class 'str'>\n", 2: 'last_name:\n Empty Value\n'}, 
       'MissingField': {'city': 'city', 'postal_code': 'postal_code'}, 'errorCount': 4}),

     ({"invoice_id": {"func": [], "type": int},
      "customer_id": {"func": [], "type": int},
      "invoice_date": {"func": [], "type": datetime.datetime}},
      {1: {'invoice_id': 1, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 4, 18, 0, 0)}, 
       2: {'invoice_id': 2, 'customer_id': 2, 'invoice_date': datetime.datetime(2020, 5, 19, 0, 0)}, 
       3: {'invoice_id': 3, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 5, 19, 0, 0)}},
      {'rowErrorMsg': {}, 'MissingField': {}, 'errorCount': 0}),       
    ],                                
)

def test_validateDictionary(fields,dictionary,expected):                

    errDict = validateDictionary(dictionary, fields)  
    print(errDict)          
    assert errDict == expected

@pytest.mark.parametrize(
    "dictionary,expected,printout",        
    [({1: {'product_id': 1, 'name': 'Pen', 'description': 'Ball Pointed, Black Ink', 'unit_price': 3}, 
       2: {'product_id': 2, 'name': 'Pencil', 'description': 'Mechanical, 0.3mm', 'unit_price': 5.0}, 
       3: {'product_id': 3, 'name': 'Eraser', 'description': 'White', 'unit_price': 2}},
      {'rowErrorMsg': {}, 'MissingField': {}, 'errorCount': 0},
      []),

     ({1: {'product_id': 1, 'name': None, 'description': 'Ball Pointed, Black Ink'}, 
       2: {'product_id': 2, 'name': 'Pencil', 'description': 'Mechanical, 0.3mm'}, 
       3: {'product_id': "A", 'name': None, 'description': 'White'}},
      {'rowErrorMsg': {1: 'name:\n Empty Value\n', 3: 'name:\n Empty Value\n'}, 'MissingField': {'unit_price': 'unit_price'}, 'errorCount': 4},
      ["MissingField:", "Row", "Empty Value", "Error Count:"]),      
    ],                                    
)

def test_printError(dictionary,expected,printout,capsys):

    fields = {'product_id': {"func": [], "type": int}, 
              'name': {"func": [], "type": str}, 
              'description': {"func": [], "type": str}, 
              'unit_price': {"func": [], "type": float}} 

    errDict = validateDictionary(dictionary, fields)
    printError(errDict)    
    captured = capsys.readouterr()

    assert errDict == expected
    for value in printout:
        assert value in captured.out

def test_inner_join():

    Ldict = {1: {'invoice_id': 1, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 4, 18, 0, 0)}, 
             2: {'invoice_id': 2, 'customer_id': 2, 'invoice_date': datetime.datetime(2020, 5, 19, 0, 0)}, 
             3: {'invoice_id': 3, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 5, 19, 0, 0)}}

    Rdict = {1: {'customer_id': 1, 'first_name': 'John', 'last_name': 'Doe', 'phone': 4031234567, 'address': '123 Fake Street', 'city': 'Calgary', 'province': 'AB', 'postal_code': 'T1X1N1'}, 
             2: {'customer_id': 2, 'first_name': 'Jane', 'last_name': 'Smith', 'phone': 7081234567, 'address': '456 Fake Avenue', 'city': 'Edmonton', 'province': 'AB', 'postal_code': 'D1Z1X1'}, 
             3: {'customer_id': 3, 'first_name': 'Noname', 'last_name': 'OK', 'phone': 7081234567, 'address': '456 Fake Avenue', 'city': 'Edmonton', 'province': 'AB', 'postal_code': 'D1Z1X1'}}

    expected = {1: {'invoice_id': 1, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 4, 18, 0, 0), 'first_name': 'John', 'last_name': 'Doe', 'phone': 4031234567, 'address': '123 Fake Street', 'city': 'Calgary', 'province': 'AB', 'postal_code': 'T1X1N1'}, 
                2: {'invoice_id': 2, 'customer_id': 2, 'invoice_date': datetime.datetime(2020, 5, 19, 0, 0), 'first_name': 'Jane', 'last_name': 'Smith', 'phone': 7081234567, 'address': '456 Fake Avenue', 'city': 'Edmonton', 'province': 'AB', 'postal_code': 'D1Z1X1'},
                3: {'invoice_id': 3, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 5, 19, 0, 0), 'first_name': 'John', 'last_name': 'Doe', 'phone': 4031234567, 'address': '123 Fake Street', 'city': 'Calgary', 'province': 'AB', 'postal_code': 'T1X1N1'}}

    result = inner_join(Ldict, Rdict, "customer_id")                

    assert expected == result

    expected = {1: {'invoice_id': 1, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 4, 18, 0, 0), 'first_name': 'John', 'last_name': 'Doe', 'phone': 4031234567, 'address': '123 Fake Street', 'city': 'Calgary', 'province': 'AB', 'postal_code': 'T1X1N1'}}

    result = inner_join(Ldict, Rdict, "customer_id", "invoice_id", 1)

    assert expected == result

    Ldict = {1: {'invoice_id': 1, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 4, 18, 0, 0)}, 
             2: {'invoice_id': 2, 'customer_id': 2, 'invoice_date': datetime.datetime(2020, 5, 19, 0, 0)}, 
             3: {'invoice_id': 3, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 5, 19, 0, 0)}}

    Rdict = {1: {'invoice_line_Item_id': 1, 'invoice_id': 1, 'product_id': 1, 'item_ref': 'Item 1', 'quantity': 3}, 
             2: {'invoice_line_Item_id': 2, 'invoice_id': 1, 'product_id': 2, 'item_ref': 'Item 2', 'quantity': 1}, 
             3: {'invoice_line_Item_id': 3, 'invoice_id': 1, 'product_id': 3, 'item_ref': 'Item 3', 'quantity': 1}, 
             4: {'invoice_line_Item_id': 4, 'invoice_id': 2, 'product_id': 2, 'item_ref': 'Item 1', 'quantity': 3}, 
             5: {'invoice_line_Item_id': 5, 'invoice_id': 2, 'product_id': 3, 'item_ref': 'Item 2', 'quantity': 2}, 
             6: {'invoice_line_Item_id': 6, 'invoice_id': 3, 'product_id': 2, 'item_ref': 'Item 1', 'quantity': 4}}             

    expected = {1: {'invoice_id': 1, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 4, 18, 0, 0), 'invoice_line_Item_id': 1, 'product_id': 1, 'item_ref': 'Item 1', 'quantity': 3}, 
                2: {'invoice_id': 1, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 4, 18, 0, 0), 'invoice_line_Item_id': 2, 'product_id': 2, 'item_ref': 'Item 2', 'quantity': 1}, 
                3: {'invoice_id': 1, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 4, 18, 0, 0), 'invoice_line_Item_id': 3, 'product_id': 3, 'item_ref': 'Item 3', 'quantity': 1}, 
                4: {'invoice_id': 2, 'customer_id': 2, 'invoice_date': datetime.datetime(2020, 5, 19, 0, 0), 'invoice_line_Item_id': 4, 'product_id': 2, 'item_ref': 'Item 1', 'quantity': 3}, 
                5: {'invoice_id': 2, 'customer_id': 2, 'invoice_date': datetime.datetime(2020, 5, 19, 0, 0), 'invoice_line_Item_id': 5, 'product_id': 3, 'item_ref': 'Item 2', 'quantity': 2}, 
                6: {'invoice_id': 3, 'customer_id': 1, 'invoice_date': datetime.datetime(2020, 5, 19, 0, 0), 'invoice_line_Item_id': 6, 'product_id': 2, 'item_ref': 'Item 1', 'quantity': 4}}

    result = inner_join(Ldict, Rdict, "invoice_id")

    assert expected == result

def fake_input(the_prompt):
    prompt_to_return_val = {
        'Enter folder path for Excel Template or type (N) to use default path: ': 'N',
        'Enter invoice template name or type (N) to use default filename: ': 'N',
        'Enter an invoice number for printing: ': 1,
    }
    val = prompt_to_return_val[the_prompt]
    return val

def test_createInvoice(monkeypatch, dirpath, outfilePrefix, infile):

    # https://docs.pytest.org/en/latest/reference.html?highlight=monkeypatch#_pytest.monkeypatch.MonkeyPatch.undo
    # https://docs.python.org/3/library/builtins.html
    # https://stackoverflow.com/questions/56498487/how-can-i-test-a-loop-with-multiple-input-calls
    with monkeypatch.context() as m:
        m.setattr('builtins.input', fake_input)               
        createInvoice(dirpath, infile)

    expected_file = os.path.join(dirpath, f"{outfilePrefix} 1.xlsx")
    assert os.path.isfile(expected_file) == True

def test_merge_input_template(monkeypatch, dirpath, infile, capsys):

    with monkeypatch.context() as m:
        m.setattr('builtins.input', lambda x: "N")                       
        status = merge_input_template(dirpath, infile)
        captured = capsys.readouterr()
    
    if (status["Validation"] == True):
        assert len(status["errMsg"]) == 0
        assert os.path.isfile(os.path.join(dirpath, infile)) == True
    else:
        for msg in status["errMsg"]:
            assert "contains error, check file" in msg

    assert "Merging: " in captured.out
    assert "Loading worksheets: " in captured.out