from openpyxl import Workbook, load_workbook
import datetime
from pathlib import Path
import os

def createInvoice(directory, file):
    dirpath = input("Enter folder path for Excel Template or type (N) to use default path: ")
    if (dirpath.casefold() == "N".casefold()):
        dirpath = directory
        print("Use default path: ", dirpath)
    infile = input("Enter invoice template name or type (N) to use default filename: ")
    if (infile.casefold() == "N".casefold()):
        infile = file
        print("Use default input file: ", infile)

    # Load and validate input Excel template
    wb = validateInvInput(dirpath, infile)
    if wb["Validation"] == False:
        print("Error in Loading, check input file.")
    else:
        try:   
            inv = input("Enter an invoice number for printing: ")
            invIsExist = False
            # Search for invoice number in the workbook
            for rows in wb["WB"]["Invoice"].values():
                if int(inv) in rows.values():
                    invIsExist = True

            if invIsExist:
                print(f"Invoice number: {inv} found")

                # Cross reference worsheets in workbook to generate invoice                
                query1 = inner_join(wb["WB"]["Invoice"], wb["WB"]["Customer"],"customer_id", "invoice_id", int(inv))                
                query2 = inner_join(query1, wb["WB"]["Invoice Line Item"],"invoice_id")
                query3 = inner_join(query2, wb["WB"]["Product"],"product_id")                

                generateExInv(dirpath, "Excel Invoice", query3)

            else:
                print(f"Invoice number: {inv} NOT found")
            
        except ValueError:
            print(f"Invoice number must be number")

    return

def inner_join(Ldict, Rdict, joinCol, condCol=None, condValue=None): 
    
    # Fetch matched columns from both tables
    newdict = {}
    rowindex = 1
    for Lrowdata in Ldict.values():        
        for Rrowdata in Rdict.values():            
            if(Lrowdata[joinCol] == Rrowdata[joinCol]):
                newrow = {}
                newrow.update(Lrowdata)
                newrow.update(Rrowdata)
                newdict[rowindex] = newrow        
                rowindex +=1    

    # Applied filter condition
    if ((condCol !=None) & (condValue != None)):
        condDict = {}
        rowindex = 1
        for rowdata in newdict.values():
            if(rowdata[condCol] == condValue):
                condDict[rowindex] = rowdata
                rowindex += 1
        return condDict
        
    return newdict  

def generateExInv(dirpath, outfilePrefix, data):

    create_status = False
    wb = Workbook()    

    try:
        invoice = data[1]["invoice_id"]

        path = os.path.join(dirpath, f"{outfilePrefix} {invoice}.xlsx")
        ws = wb.active

        # Generate header information
        ws['A1'] = "Invoice Date:"
        ws['B1'] = data[1]["invoice_date"]
        ws['A2'] = "Invoice Number:"
        ws['B2'] = invoice
        ws['E1'] = "Billed to:"
        ws['F1'] = f"{data[1]['first_name']} {data[1]['last_name']}"
        ws['F2'] = data[1]["phone"]
        ws['F3'] = data[1]["address"]
        ws['F4'] = f"{data[1]['city']}, {data[1]['province']}"
        ws['F5'] = data[1]["postal_code"]
        ws['B8'] = "Name"
        ws['C8'] = "Description"
        ws['D8'] = "Qty"
        ws['E8'] = "Price"
        ws['F8'] = "Amount"

        # Generate line items
        offset = 8
        amt = {}
        for row, rowdata in data.items():
            ws[f"A{row+offset}"] = rowdata["item_ref"]
            ws[f"B{row+offset}"] = rowdata["name"]
            ws[f"C{row+offset}"] = rowdata["description"]
            ws[f"D{row+offset}"] = rowdata["quantity"]
            ws[f"E{row+offset}"] = rowdata["unit_price"]
            ws[f"F{row+offset}"] = rowdata["unit_price"] * rowdata["quantity"]
            amt[row] = rowdata["unit_price"] * rowdata["quantity"]

        offset = offset + len(data) + 2        
        subtotal = sum(amt.values())
        gst = subtotal*0.05
        total = subtotal + gst
        ws[f"E{offset}"] = "Subtotal"
        ws[f"F{offset}"] = subtotal
        ws[f"E{offset+1}"] = "GST"
        ws[f"F{offset+1}"] = gst
        ws[f"E{offset+2}"] = "Total"
        ws[f"F{offset+2}"] = total

        wb.save(path)
        create_status = True
    except: # catch *all* exceptions
        pass

    return create_status

# Exercise - Validate and Merge the data

# Validation
# Write python code and use the OpenPyXL package to validate that your input worksheet data is in 
# the right format and reasonable for your system. Validate all the other group’s data as well.  
# Larry will come and modify your input spreadsheet to try and break your system.  
# Your validation code should catch these problems.

# Think about how to validate your inputs from ANY source, i.e. Excel sheets, tables, etc.  
# You may not always be loading data from this spreadsheet.  Try to build your validation code 
# to handle ANY input format.
def validateInvInput(dirpath, filename):

    valid_status = True  
    missingSheets = []
    wsNames = ["Customer", "Invoice", "Invoice Line Item", "Product"]
    wsfields = getFields()
    dictionary = {}
    errDict = {}
    wbDict = {}

    try:
        path = os.path.join(dirpath, filename)
        wb = load_workbook(path)

        for wsname in wsNames:
            if not wsname in wb.sheetnames:
                valid_status &= False
                missingSheets.append(wsname)            

        if(valid_status == False):
            for wsname in missingSheets:
                msg = f"{wsname} ,"
            msg = msg.strip(" ,")
            print(f"Missing worksheets: {msg}")
        else:
            # No workbook error, load and verify each worksheet            
            for wsname in wsNames:
                print("===================================================")
                print(f"Loading worksheets: ", wsname)
                dictionary[wsname] = loadWStoDictionary(wb[wsname], wsfields[wsname].keys())                
                errDict[wsname] = validateDictionary(dictionary[wsname], wsfields[wsname])                                
                if (errDict[wsname]["errorCount"]):
                    printError(errDict[wsname])
                    valid_status = False                    
            print("===================================================")
                                                
    except: # catch *all* exceptions
        valid_status = False
    
    wbDict["WB"] = dictionary
    wbDict["Error"] = errDict
    wbDict["Validation"] = valid_status

    return wbDict

# Load worksheet into a nested dictionary
#   Result dictionary format:
#       {Row1Key: {Col1Key: value1, Col2Key: value2}
#       {Row2Key: {Col1Key: value3, Col2Key: value4}
def loadWStoDictionary(ws, fields):
    titleIndex = {}
    dictionary = {}
    buffer_dict = {}

    # Map the title index
    for field in fields:        
        col_index = 1
        for col in ws[1]:
            if col.value == field:                
                titleIndex[col_index] = field
            col_index += 1

    # Load the data into dictionary
    row_index = 1    
    for row in ws.values:
        col_index = 1        
        for value in row:            
            if((col_index in titleIndex.keys()) & (row_index!=1)):
                colkey = titleIndex[col_index]                
                buffer_dict[colkey] = value
            col_index += 1

        if (row_index!=1):
            rowkey = row_index - 1
            dictionary[rowkey] = buffer_dict
            buffer_dict = {}

        row_index += 1

    return dictionary

def validateDictionary(dictionary, fields):
    missingField = {}
    errDict = {}
    errRow = {}
    errCount = 0
          
    # Loop through all data rows    
    for rownum, rowdata in dictionary.items():
        for field, funcType in fields.items():
            if not(field in rowdata.keys()):
                # Missing field => skip the rest of checks                
                if not(field in missingField.keys()):
                    missingField[field] = field
                    errCount += 1
            else:                
                # No missing field => perform other value check
                # Loop through all columns in a row                 
                msg = ""       
                for col_title, value in rowdata.items():                                                  
                    ##################### Need a unique value check for certain columns too                   
                    if (col_title == field):                                            
                        if (value == None):
                            msg = "Empty Value\n"                            
                        else:   
                            # Validate basic data type
                            msg = checkType(value, funcType["type"])

                            # Validate conditions applied to specific title
                            for func in funcType["func"]:
                                msg = func(value)

                        # Only save message if there is error, 
                        # only count one error per cell value regardless 
                        # how many violations in one value
                        if (len(msg) > 0):
                            errCount += 1
                            errRow[rownum] = f"{field}:\n {msg}"                                    

    errDict["rowErrorMsg"] = errRow
    errDict["MissingField"] = missingField
    errDict["errorCount"] = errCount    
        
    return errDict

def checkType(value, expected_type):    

    if (type(value) is expected_type):
        return ""
    elif (expected_type == float):
        if (type(value) is int):
            return ""
        else:
            return f"Incorrect Data Type, Expected: {expected_type}, Received: {type(value)}\n"
    else:
        return f"Incorrect Data Type, Expected: {expected_type}, Received: {type(value)}\n"

def checkPostalCode(value):
    postalcode = value.replace(" ", "")
    errMsg = ""

    if not (len(postalcode) == 6):
        errMsg = "Incorrect Postal Format\n"
    else:
        if not (postalcode[0].isalpha()):
            errMsg = "Incorrect Postal Format\n"
        elif not (postalcode[1].isdigit()):
            errMsg = "Incorrect Postal Format\n"
        elif not (postalcode[2].isalpha()):
            errMsg = "Incorrect Postal Format\n"
        elif not (postalcode[3].isdigit()):
            errMsg = "Incorrect Postal Format\n"
        elif not (postalcode[4].isalpha()):
            errMsg = "Incorrect Postal Format\n"
        elif not (postalcode[5].isdigit()):
            errMsg = "Incorrect Postal Format\n"                                

    return errMsg       

def checkPhone(value):
    errMsg = ""
    phone = str(value).replace(" ", "")
    phone = phone.replace("-", "")
    phone = phone.replace("(", "")
    phone = phone.replace(")", "")
    if not (phone.isdigit()):
        errMsg = "Incorrect Phone Format\n"
    elif not (len(phone) == 10):
        errMsg = "Incorrect Phone Format\n"    

    return errMsg

def getFields():
    wbFields = {}

    fields = {"customer_id": {"func": [], "type": int},
              "first_name": {"func": [], "type": str},
              "last_name": {"func": [], "type": str},
              "phone": {"func": [checkPhone], "type": str},
              "address": {"func": [], "type": str},
              "city": {"func": [], "type": str},
              "province": {"func": [], "type": str},
              "postal_code": {"func": [checkPostalCode], "type": str}}

    wbFields["Customer"] = fields

    fields = {"invoice_id": {"func": [], "type": int},
              "customer_id": {"func": [], "type": int},
              "invoice_date": {"func": [], "type": datetime.datetime}}

    wbFields["Invoice"] = fields              

    fields = {"invoice_line_Item_id": {"func": [], "type": int},
              "invoice_id": {"func": [], "type": int},
              "product_id": {"func": [], "type": int},
              "item_ref": {"func": [], "type": str},
              "quantity": {"func": [], "type": int}}

    wbFields["Invoice Line Item"] = fields              

    fields = {'product_id': {"func": [], "type": int}, 
              'name': {"func": [], "type": str}, 
              'description': {"func": [], "type": str}, 
              'unit_price': {"func": [], "type": float}} 
    
    wbFields["Product"] = fields

    return wbFields

def printError(errMsg):
    if (errMsg["errorCount"] > 0):
        if (errMsg["MissingField"]):
            print(f"MissingField:")                        
            for key, value in errMsg["MissingField"].items():
                print(f"   {key}\n")

        if (len(errMsg["rowErrorMsg"]) > 0):
            for key, value in errMsg["rowErrorMsg"].items():
                print(f"Row {key}:\n {value}")

        print(f"Error Count: {errMsg['errorCount']}")

# Merge
# Write python code to merge all the other groups’ data into a single well-formatted workbook 
# that matches your existing data model. If you are the first group, create a second one and merge it.  
# Validate the new workbook and create some invoices.
def merge_input_template(directory, file):

    valid_merge_status = True
    errmsg = []

    dirpath = input("Enter folder path for Excel Template or type (N) to use default path: ")
    if (dirpath.casefold() == "N".casefold()):
        dirpath = directory
        print("Use default path: ", dirpath)

    # Remove the previous merged template if exists
    templatefile = os.path.join(dirpath, file)
    if (os.path.isfile(templatefile)):
        os.remove(templatefile)

    for filename in os.listdir(dirpath):
        if filename.endswith(".xlsx"):
            print(f"================== Merging: {filename} =======================")
            dictionary = validateInvInput(dirpath, filename)
            if dictionary["Validation"] == True:
                valid_merge_status = valid_merge_status & True
                
                create_template(dirpath, "Merge_template.xlsx", dictionary["WB"])
            else:
                msg = F"{filename} contains error, check file"
                errmsg.append(msg)
                print(msg)
                valid_merge_status = valid_merge_status & False
    
    return {"errMsg": errmsg, "Validation": valid_merge_status}

def create_template(dirpath, filename, dictionary):
    filename = os.path.join(dirpath, filename)
    if (os.path.isfile(filename)):        
        wb = load_workbook(filename)

        # Loop through worksheet
        for wsname, ws_dict in dictionary.items():
            ws = wb[wsname]
            row_offset = len(ws["A"])
                        
            # Loop through rows
            for rownum, rowdata in ws_dict.items():
                # Loop through columns                                
                colindex = 1
                for value in rowdata.values():
                    ws.cell(row=rownum+row_offset, column=colindex).value = value
                    colindex += 1

        wb.save(filename)                
          
    else:            
        wb = Workbook()
        row_offset = 1

        # Loop through worksheet
        for wsname, ws_dict in dictionary.items():
            ws = wb.create_sheet(wsname)

            # # Create title
            # ws.cell(row=1, column=col).value = col_title

            # Loop through rows
            for rownum, rowdata in ws_dict.items():
                # Loop through columns
                colindex = 1                
                for col_title, value in rowdata.items():
                    # Create title
                    ws.cell(row=1, column=colindex).value = col_title

                    # Create non-title rows
                    ws.cell(row=rownum+row_offset, column=colindex).value = value
                    colindex += 1

        ws = wb['Sheet']
        wb.remove(ws)
        wb.save(filename)

    return 

# Main program code to execute the functions
if __name__ == '__main__':
    print("--- Starting", __file__)
    defaultPath = Path("C:/code/cohort4/python-Excel/template")
    defaultTemplatefile = os.path.join(defaultPath, "Merge_template.xlsx")

    merge_input_template(defaultPath, defaultTemplatefile)
    createInvoice(defaultPath, defaultTemplatefile)
