import pytest
import os
import datetime
from pathlib import Path
from openpyxl import load_workbook
from testfixtures import TempDirectory
from generateExInv import generateExInv, validateInvInput, loadWStoDictionary, validateDictionary, checkType, checkPostalCode, checkPhone, printError, createInvoice, inner_join, merge_input_template

@pytest.fixture
def dirpath():
    return Path("C:/code/cohort4/python-Excel/template")
@pytest.fixture    
def outfilePrefix():
    return "Excel Invoice"
@pytest.fixture    
def infile():
    return "Merge_template.xlsx"

@pytest.fixture(params=[
        {"fields": ["customer_id", "first_name", "last_name", "phone", "address", "city", "province", "postal_code"],
         "wsname": "Customer"},
        {"fields": ["invoice_id", "customer_id", "invoice_date"],
         "wsname": "Invoice"},         
        {"fields": ["invoice_line_Item_id", "invoice_id", "product_id", "item_ref", "quantity"],
         "wsname": "Invoice Line Item"},                  
        {"fields": ["product_id", "name", "description", "unit_price"],
         "wsname": "Product"},                           
    ])

def setup(request):
    retVal = request.param
    # print("\nSetup! retVal = {}".format(retVal))
    return retVal

#################################################################
# Exercise - Populate the data
#################################################################    
def test_generateExInv(dirpath, outfilePrefix):    

    data = {'invoice_id': {1: 1, 2: 1, 3: 1}, 'customer_id': {1: 1, 2: 1, 3: 1}, 
    'invoice_date': {1: datetime.datetime(2020, 4, 18, 0, 0), 2: datetime.datetime(2020, 4, 18, 0, 0), 3: datetime.datetime(2020, 4, 18, 0, 0)}, 
    'first_name': {1: 'John', 2: 'John', 3: 'John'}, 'last_name': {1: 'Doe', 2: 'Doe', 3: 'Doe'}, 
    'phone': {1: 4031234567, 2: 4031234567, 3: 4031234567}, 
    'address': {1: '123 Fake Street', 2: '123 Fake Street', 3: '123 Fake Street'}, 
    'city': {1: 'Calgary', 2: 'Calgary', 3: 'Calgary'}, 
    'province': {1: 'AB', 2: 'AB', 3: 'AB'}, 
    'postal_code': {1: 'T1X1N1', 2: 'T1X1N1', 3: 'T1X1N1'}, 
    'invoice_line_Item_id': {1: 1, 2: 2, 3: 3}, 'product_id': {1: 1, 2: 2, 3: 3}, 
    'item_ref': {1: 'Item 1', 2: 'Item 2', 3: 'Item 3'}, 
    'quantity': {1: 3, 2: 1, 3: 1}, 
    'name': {1: 'Pen', 2: 'Pencil', 3: 'Eraser'}, 
    'description': {1: 'Ball Pointed, Black Ink', 2: 'Mechanical, 0.3mm', 3: 'White'}, 
    'unit_price': {1: 3, 2: 5, 3: 2}}

    outfile = f"{outfilePrefix} 1.xlsx"

    with TempDirectory() as d:      
        create_status = generateExInv(d.path,outfilePrefix,data)        
        filename = os.path.join(d.path, outfile)
        existEX = os.path.exists(filename)
        wb = load_workbook(filename)
        ws = wb.active       
        d.listdir()      

    assert existEX == True
    assert create_status == True
    assert ws['A1'].value == "Invoice Date:"
    assert ws['B1'].value == datetime.datetime(2020, 4, 18, 0, 0)
    assert ws['A2'].value == "Invoice Number:"  
    assert ws['B2'].value == 1
    assert ws['E1'].value == "Billed to:"
    assert ws['F1'].value == "John Doe"

    assert ws['A9'].value == "Item 1"
    assert ws['B9'].value == "Pen"
    assert ws['C9'].value == "Ball Pointed, Black Ink"

#################################################################
# Exercise - Validate the data
#################################################################
def test_validateInvInput(monkeypatch, capsys, dirpath, infile):

    filename = os.path.join(dirpath, infile)
    
    ws1_exist = False
    ws2_exist = False
    ws3_exist = False
    ws4_exist = False    

    input_exist = os.path.isfile(filename)
    if input_exist:
        wb = load_workbook(filename)
        if "Customer" in wb.sheetnames:
            ws1_exist = True
        if "Invoice" in wb.sheetnames:
            ws2_exist = True 
        if "Invoice Line Item" in wb.sheetnames:        
            ws3_exist = True
        if "Product" in wb.sheetnames:
            ws4_exist = True               

    wbDict = validateInvInput(dirpath, infile)
    valid_status = wbDict["Validation"]
    missing_ws = ws1_exist & ws2_exist & ws3_exist & ws4_exist
    
    captured = capsys.readouterr()

    assert "WB" in wbDict.keys()
    assert "Error" in wbDict.keys()
    assert "Validation" in wbDict.keys()
    assert "Customer" in wbDict["WB"]
    assert "Invoice" in wbDict["WB"]
    assert "Invoice Line Item" in wbDict["WB"]
    assert "Product" in wbDict["WB"]

    if (valid_status == False):
        assert "Error Count: " in captured.out
    if (missing_ws == False):
        assert "Missing worksheets: " in captured.out
        if (ws1_exist == False):
            assert "Customer" in captured.out
        if (ws2_exist == False):
            assert "Invoice" in captured.out    
        if (ws3_exist == False):
            assert "Invoice Line Item" in captured.out
        if (ws4_exist == False):
            assert "Product" in captured.out

def test_loadWStoDictionary(dirpath, infile, setup):
    
    filename = os.path.join(dirpath, infile)
    input_exist = os.path.isfile(filename)
    if input_exist:
        wb = load_workbook(filename) 
        ws = wb[setup["wsname"]]
        dictionary = loadWStoDictionary(ws, setup["fields"])
        print(dictionary)

        for field in setup["fields"]:
            assert field in dictionary.keys()

def test_checkType():    
    assert checkType(123, int) == ""                            
    assert checkType("123", int) == "Incorrect Data Type, Expected: <class 'int'>, Received: <class 'str'>\n"
    assert checkType("123", str) == ""                            
    assert checkType("ABC", str) == ""                            
    assert checkType("ABC", int) == "Incorrect Data Type, Expected: <class 'int'>, Received: <class 'str'>\n"
    assert checkType(12.3, float) == ""
    assert checkType(12, float) == ""
    assert checkType(12.0, float) == ""
    assert checkType("12.3", float) == "Incorrect Data Type, Expected: <class 'float'>, Received: <class 'str'>\n"

def test_checkPostalCode():    
    assert checkPostalCode("123") == "Incorrect Postal Format"  
    assert checkPostalCode("123567") == "Incorrect Postal Format"  
    assert checkPostalCode("1a3a6a") == "Incorrect Postal Format"  
    assert checkPostalCode("a2a5a7") == ""  
    assert checkPostalCode("a2a 5a7") == ""

def test_checkPhone():
    assert checkPhone("4031234567") == ""      
    assert checkPhone("403 -123 -4567") == ""
    assert checkPhone("(403)123-4567") == ""
    assert checkPhone("(403)123-45678") == "Incorrect Phone Format"
    assert checkPhone("(403)123-45A7") == "Incorrect Phone Format"
    assert checkPhone("(403)abc-45A7") == "Incorrect Phone Format"

@pytest.mark.parametrize(
    "fields,dictionary,expected",
    [({"customer_id": {"func": [], "type": int},
       "first_name": {"func": [], "type": str},
       "last_name": {"func": [], "type": str},
       "phone": {"func": [checkPhone], "type": str},
       "address": {"func": [], "type": str},
       "city": {"func": [], "type": str},
       "province": {"func": [], "type": str},
       "postal_code": {"func": [checkPostalCode], "type": str}},
      {'customer_id': {1: 1, 2: 2, 3: 3}, 
       'first_name': {1: 'John', 2: 'Jane', 3: 'Noname'}, 
       'last_name': {1: 'Doe', 2: 'Smith', 3: None}, 
       'phone': {1: 4031234567, 2: 7081234567, 3: 9051234567}, 
       'address': {1: '123 Fake Street', 2: '456 Fake Avenue', 3: '456 Test Dr'}, 
       'city': {1: 'Calgary', 2: 'Edmonton', 3: 'Halifax'}, 
       'province': {1: 'AB', 2: 'AB', 3: 'NS'}, 
       'postal_code': {1: 'T1X1N1', 2: 'D1Z1X1', 3: 'A1B1C1'}},        
      {'customer_id': {}, 'first_name': {}, 'last_name': {3: 'Empty Value\n'}, 
       'phone': {}, 'address': {}, 'city': {}, 'province': {}, 'postal_code': {}, 
       'MissingField': {}, 'errorCount': 1}),

     ({"customer_id": {"func": [], "type": int},
       "first_name": {"func": [], "type": str},
       "last_name": {"func": [], "type": str},
       "phone": {"func": [checkPhone], "type": str},
       "address": {"func": [], "type": str},
       "city": {"func": [], "type": str},
       "province": {"func": [], "type": str},
       "postal_code": {"func": [checkPostalCode], "type": str}},
      {'customer_id': {1: 1, 2: 2, 3: 3}, 
       'first_name': {1: 'John', 2: 'Jane', 3: 'Noname'}, 
       'last_name': {1: 'Doe', 2: 'Smith', 3: None}, 
       'phone': {1: "403AB34567", 2: 7081234567, 3: 9051234567}, 
       'address': {1: '123 Fake Street', 2: '456 Fake Avenue', 3: '456 Test Dr'},  
       'province': {1: 'AB', 2: 'AB', 3: 'NS'}},
      {'customer_id': {}, 'first_name': {}, 'last_name': {3: 'Empty Value\n'}, 
       'phone': {1: 'Incorrect Phone Format\n'}, 'address': {}, 'province': {}, 
       'MissingField': {'city': 'city', 'postal_code': 'postal_code'}, 'errorCount': 4}),

     ({"invoice_id": {"func": [], "type": int},
      "customer_id": {"func": [], "type": int},
      "invoice_date": {"func": [], "type": datetime.datetime}},
     {'invoice_id': {1: 1, 2: 2, 3: 3}, 
      'customer_id': {1: 1, 2: 2, 3: 1}, 
      'invoice_date': {1: datetime.datetime(2020, 4, 18, 0, 0), 
                       2: datetime.datetime(2020, 5, 19, 0, 0), 
                       3: datetime.datetime(2020, 5, 19, 0, 0)}},
     {'invoice_id': {}, 'customer_id': {}, 'invoice_date': {}, 'MissingField': {}, 'errorCount': 0}),

     ({"invoice_line_Item_id": {"func": [], "type": int},
       "invoice_id": {"func": [], "type": int},
       "product_id": {"func": [], "type": int},
       "item_ref": {"func": [], "type": str},
       "quantity": {"func": [], "type": int}},
      {'invoice_line_Item_id': {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6}, 
       'invoice_id': {1: 1, 2: 1, 3: 1, 4: 2, 5: 2, 6: 3}, 
       'product_id': {1: 1, 2: 2, 3: 3, 4: 2, 5: 3, 6: 2}, 
       'item_ref': {1: 'Item 1', 2: 'Item 2', 3: 'Item 3', 4: 'Item 1', 5: 'Item 2', 6: 'Item 1'}, 
       'quantity': {1: 3, 2: 1, 3: 1, 4: 3, 5: 2, 6: 4}},
      {'invoice_line_Item_id': {}, 'invoice_id': {}, 'product_id': {}, 'item_ref': {}, 'quantity': {},
       'MissingField': {}, 'errorCount': 0}),

     ({'product_id': {"func": [], "type": int}, 
       'name': {"func": [], "type": str}, 
       'description': {"func": [], "type": str}, 
       'unit_price': {"func": [], "type": float}},                  
      {'product_id': {1: 1, 2: 2, 3: 3}, 
       'name': {1: 'Pen', 2: 'Pencil', 3: 'Eraser'}, 
       'description': {1: 'Ball Pointed, Black Ink', 2: 'Mechanical, 0.3mm', 3: 'White'}, 
       'unit_price': {1: 3, 2: 5.0, 3: 2}},
      {'product_id': {}, 'name': {}, 'description': {}, 'unit_price': {}, 'MissingField': {}, 'errorCount': 0})
    ],                                
)

def test_validateDictionary(fields,dictionary,expected):                

    errDict = validateDictionary(dictionary, fields)            
    assert errDict == expected

@pytest.mark.parametrize(
    "dictionary,expected,printout",        
    [({'product_id': {1: 1, 2: 2, 3: 3}, 
       'name': {1: 'Pen', 2: 'Pencil', 3: 'Eraser'}, 
       'description': {1: 'Ball Pointed, Black Ink', 2: 'Mechanical, 0.3mm', 3: 'White'}, 
       'unit_price': {1: 3, 2: 5.0, 3: 2}},
      {'product_id': {}, 'name': {}, 'description': {}, 'unit_price': {}, 'MissingField': {}, 'errorCount': 0},
      []),

     ({'product_id': {1: 1, 2: 2, 3: "A"}, 
       'name': {1: None, 2: 'Pencil', 3: None}, 
       'description': {1: 'Ball Pointed, Black Ink', 2: 'Mechanical, 0.3mm', 3: 'White'},},
      {'product_id': {3: "Incorrect Data Type, Expected: <class 'int'>, Received: <class 'str'>\n"}, 
       'name': {1: 'Empty Value\n', 3: 'Empty Value\n'}, 'description': {}, 
       'MissingField': {'unit_price': 'unit_price'}, 'errorCount': 4},
       ["Error in column", "Row: ", "Empty Value", "MissingField", "Error Count:"]),      
    ],                                    
)

def test_printError(dictionary,expected,printout,capsys):

    fields = {'product_id': {"func": [], "type": int}, 
              'name': {"func": [], "type": str}, 
              'description': {"func": [], "type": str}, 
              'unit_price': {"func": [], "type": float}} 

    errDict = validateDictionary(dictionary, fields) 
    printError(errDict)    
    captured = capsys.readouterr()

    assert errDict == expected
    for value in printout:
        assert value in captured.out

def test_inner_join():
    Ldict = {'invoice_id': {1: 1, 2: 2, 3: 3}, 'customer_id': {1: 1, 2: 2, 3: 1}, 
             'invoice_date': {1: datetime.datetime(2020, 4, 18, 0, 0), 2: datetime.datetime(2020, 5, 19, 0, 0), 3: datetime.datetime(2020, 5, 19, 0, 0)}}
    Rdict = {'customer_id': {1: 1, 2: 2, 3: 3}, 'first_name': {1: 'John', 2: 'Jane', 3: 'Noname'}, 
             'last_name': {1: 'Doe', 2: 'Smith', 3:'OK'}, 'phone': {1: 4031234567, 2: 7081234567, 3: 7081234567}, 
             'address': {1: '123 Fake Street', 2: '456 Fake Avenue', 3: '456 Fake Avenue'}, 
             'city': {1: 'Calgary', 2: 'Edmonton', 3: 'Edmonton'}, 'province': {1: 'AB', 2: 'AB', 3: 'AB'}, 
             'postal_code': {1: 'T1X1N1', 2: 'D1Z1X1', 3: 'D1Z1X1'}}    

    expected = {'invoice_id': {1: 1, 2: 2, 3: 3}, 'customer_id': {1: 1, 2: 2, 3: 1}, 
                'invoice_date': {1: datetime.datetime(2020, 4, 18, 0, 0), 2: datetime.datetime(2020, 5, 19, 0, 0), 3: datetime.datetime(2020, 5, 19, 0, 0)}, 
                'first_name': {1: 'John', 2: 'Jane', 3: 'John'}, 'last_name': {1: 'Doe', 2: 'Smith', 3: 'Doe'}, 
                'phone': {1: 4031234567, 2: 7081234567, 3: 4031234567}, 'address': {1: '123 Fake Street', 2: '456 Fake Avenue', 3: '123 Fake Street'}, 
                'city': {1: 'Calgary', 2: 'Edmonton', 3: 'Calgary'}, 'province': {1: 'AB', 2: 'AB', 3: 'AB'}, 
                'postal_code': {1: 'T1X1N1', 2: 'D1Z1X1', 3: 'T1X1N1'}}

    result = inner_join(Ldict, Rdict, "customer_id")                

    assert expected == result

    expected = {'invoice_id': {1: 1}, 'customer_id': {1: 1}, 'invoice_date': {1: datetime.datetime(2020, 4, 18, 0, 0)}, 'first_name': {1: 'John'},
                'last_name': {1: 'Doe'}, 'phone': {1: 4031234567}, 'address': {1: '123 Fake Street'}, 
                'city': {1: 'Calgary'}, 'province': {1: 'AB'}, 'postal_code': {1: 'T1X1N1'}}

    result = inner_join(Ldict, Rdict, "customer_id", "invoice_id", 1)

    assert expected == result

    Ldict = {'invoice_id': {1: 1}, 'customer_id': {1: 1}, 'invoice_date': {1: datetime.datetime(2020, 4, 18, 0, 0)}, 
             'first_name': {1: 'John'}, 'last_name': {1: 'Doe'}, 'phone': {1: 4031234567}, 
             'address': {1: '123 Fake Street'}, 'city': {1: 'Calgary'}, 'province': {1:'AB'}, 
             'postal_code': {1: 'T1X1N1'}}

    Rdict = {'invoice_line_Item_id': {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6}, 'invoice_id': {1: 1, 2: 1, 3: 1, 4: 2, 5: 2, 6: 3}, 
             'product_id': {1: 1, 2: 2, 3: 3, 4: 2, 5: 3, 6: 2}, 
             'item_ref': {1: 'Item 1', 2: 'Item 2', 3: 'Item 3', 4: 'Item 1', 5: 'Item 2', 6: 'Item 1'}, 
             'quantity': {1: 3, 2: 1, 3: 1, 4: 3, 5: 2, 6: 4}}

    expected = {'invoice_id': {1: 1, 2: 1, 3: 1}, 'customer_id': {1: 1, 2: 1, 3: 1}, 
                'invoice_date': {1: datetime.datetime(2020, 4, 18, 0, 0), 2: datetime.datetime(2020, 4, 18, 0, 0), 3: datetime.datetime(2020, 4, 18, 0, 0)}, 
                'first_name': {1: 'John', 2: 'John', 3: 'John'}, 'last_name': {1: 'Doe', 2: 'Doe', 3: 'Doe'}, 
                'phone': {1: 4031234567, 2: 4031234567, 3: 4031234567}, 
                'address': {1: '123 Fake Street', 2: '123 Fake Street', 3: '123 Fake Street'}, 
                'city': {1: 'Calgary', 2: 'Calgary', 3: 'Calgary'}, 'province': {1: 'AB', 2: 'AB', 3: 'AB'}, 
                'postal_code': {1: 'T1X1N1', 2: 'T1X1N1', 3: 'T1X1N1'}, 'invoice_line_Item_id': {1: 1, 2: 2, 3: 3}, 
                'product_id': {1: 1, 2: 2, 3: 3}, 'item_ref': {1: 'Item 1', 2: 'Item 2', 3: 'Item 3'}, 
                'quantity': {1: 3, 2: 1, 3: 1}}

    result = inner_join(Ldict, Rdict, "invoice_id")

    assert expected == result

def fake_input(the_prompt):
    prompt_to_return_val = {
        'Enter folder path for Excel Template or type (N) to use default path: ': 'N',
        'Enter invoice template name or type (N) to use default filename: ': 'N',
        'Enter an invoice number for printing: ': 1,
    }
    val = prompt_to_return_val[the_prompt]
    return val

def test_createInvoice(monkeypatch, dirpath, outfilePrefix, infile):

    # https://docs.pytest.org/en/latest/reference.html?highlight=monkeypatch#_pytest.monkeypatch.MonkeyPatch.undo
    # https://docs.python.org/3/library/builtins.html
    # https://stackoverflow.com/questions/56498487/how-can-i-test-a-loop-with-multiple-input-calls
    with monkeypatch.context() as m:
        m.setattr('builtins.input', fake_input)               
        createInvoice(dirpath, infile)

    expected_file = os.path.join(dirpath, f"{outfilePrefix} 1.xlsx")
    assert os.path.isfile(expected_file) == True

def test_merge_input_template(monkeypatch, dirpath, infile, capsys):

    with monkeypatch.context() as m:
        m.setattr('builtins.input', lambda x: "N")                       
        status = merge_input_template(dirpath, infile)
        captured = capsys.readouterr()
    
    if (status["Validation"] == True):
        assert len(status["errMsg"]) == 0
        assert os.path.isfile(os.path.join(dirpath, infile)) == True
    else:
        for msg in status["errMsg"]:
            assert "contains error, check file" in msg

    assert "Merging: " in captured.out
    assert "Loading worksheets: " in captured.out